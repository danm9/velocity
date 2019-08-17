import openpyxl
wb = openpyxl.load_workbook('C:\\Users\\dminahan\\python\\practice\\produceSales.xlsx')
sheet = wb['Sheet']

price_updates = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_updates:
        sheet.cell(row=rowNum, column=2).value = price_updates[produceName]

wb.save('C:\\Users\\dminahan\\python\\practice\\produceSales.xlsx')S